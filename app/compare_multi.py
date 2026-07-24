"""
Multi-supplier comparison.

Takes the merged inventory plus the CURRENT sheet from every supplier, and
produces one workbook where each product shows:
  - his Last Cost and Selling Price
  - a cost column per supplier (or "not carried")
  - Best Cost + Best Supplier (the cheapest that carries it)
  - margin computed against the Best Cost, with the usual flags

Matching is by Barcode throughout.
"""
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import CellIsRule

from .compare import (
    FONT, LOW_MARGIN_THRESHOLD, SUPPLIER_REQUIRED,
    CompareError, _norm_cols, _check_columns, merge_inventories,
)

NOT_CARRIED = "not carried"

BASE_COLS = [
    "Barcode", "SKU", "Brand", "Perfume Name", "Description", "Size (ml)",
    "Gender", "Category", "Stock Qty", "Country of Origin",
    "Last Cost", "Market/Selling Price",
]


def build_multi_supplier_comparison(inventory_files: list, supplier_sheets: list) -> tuple:
    """inventory_files: [(filename, bytes)]
       supplier_sheets: [(code, name, ref, bytes)]
       Returns (BytesIO xlsx, info dict).
    """
    if not supplier_sheets:
        raise CompareError("No supplier sheets are saved yet. Add a supplier and upload their sheet first.")

    shop, merge_info = merge_inventories(inventory_files)

    # Read each supplier's current sheet down to Barcode + their cost
    per_supplier = []          # (code, name, ref, DataFrame[Barcode, <code>])
    for code, name, ref, data in supplier_sheets:
        try:
            df = _norm_cols(pd.read_excel(BytesIO(data)))
        except Exception as e:
            raise CompareError(f"Could not read the sheet for {name} ({ref}) as Excel. ({e})")
        cols = _check_columns(df, SUPPLIER_REQUIRED, f"supplier sheet '{ref}'")
        slim = df[[cols["Barcode"], cols["Current Cost"]]].rename(
            columns={cols["Barcode"]: "Barcode", cols["Current Cost"]: code}
        )
        slim = slim.drop_duplicates(subset="Barcode", keep="last")
        per_supplier.append((code, name, ref, slim))

    out = shop.copy()
    codes = []
    for code, name, ref, slim in per_supplier:
        out = out.merge(slim, on="Barcode", how="left")
        codes.append(code)

    # Best cost across suppliers (ignoring those that don't carry it).
    # Products no supplier carries produce an all-NaN row, which idxmin rejects,
    # so compute the winner only for rows that have at least one price.
    cost_block = out[codes].apply(pd.to_numeric, errors="coerce")
    out["Best Cost"] = cost_block.min(axis=1, skipna=True)
    carried = out["Best Cost"].notna()
    best_code = pd.Series(pd.NA, index=out.index, dtype="object")
    if carried.any():
        best_code.loc[carried] = cost_block.loc[carried].idxmin(axis=1)
    out["Best Supplier"] = best_code
    out["Match Status"] = np.where(
        carried, "Available", "NOT carried by any supplier"
    )

    # Products available from at least one supplier first
    out["_rank"] = np.where(carried, 0, 1)
    out = out.sort_values("_rank", kind="stable").drop(columns="_rank").reset_index(drop=True)

    present_base = [c for c in BASE_COLS if c in out.columns]
    ordered = present_base + codes + ["Best Cost", "Best Supplier", "Match Status"]
    out = out[[c for c in ordered if c in out.columns]]

    info = {
        **merge_info,
        "supplier_count": len(per_supplier),
        "suppliers": [{"code": c, "name": n, "ref": r} for c, n, r, _ in per_supplier],
        "products_available": int(carried.sum()),
        "products_unavailable": int((~carried).sum()),
    }
    return _style_multi(out, codes, per_supplier), info


def _style_multi(df: pd.DataFrame, codes: list, per_supplier: list) -> BytesIO:
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    ws.title = "Comparison"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    supp_fill = PatternFill("solid", fgColor="2E6B4F")   # supplier cost columns
    calc_fill = PatternFill("solid", fgColor="C55A11")   # computed columns
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    nrow = ws.max_row
    ncol = ws.max_column
    col = {cell.value: i + 1 for i, cell in enumerate(ws[1])}

    lc_c = get_column_letter(col["Last Cost"])
    sp_c = get_column_letter(col["Market/Selling Price"])
    best_c = get_column_letter(col["Best Cost"])

    # Append computed columns
    new_headers = ["Cost Diff (AED)", "Cost Change %", "Margin", "Margin %",
                   "Price Flag", "Margin Flag", "Trade Direction"]
    start = ncol + 1
    for i, h in enumerate(new_headers):
        ws.cell(row=1, column=start + i, value=h)
    C = {h: get_column_letter(start + i) for i, h in enumerate(new_headers)}

    lm = LOW_MARGIN_THRESHOLD
    for r in range(2, nrow + 1):
        lc, sp, bc = f"{lc_c}{r}", f"{sp_c}{r}", f"{best_c}{r}"
        ws[f'{C["Cost Diff (AED)"]}{r}'] = f'=IF({bc}="","",{bc}-{lc})'
        ws[f'{C["Cost Change %"]}{r}'] = f'=IF(OR({bc}="",{lc}=0),"",({bc}-{lc})/{lc})'
        ws[f'{C["Margin"]}{r}'] = f'=IF({bc}="","",{sp}-{bc})'
        ws[f'{C["Margin %"]}{r}'] = f'=IF(OR({bc}="",{sp}=0),"",({sp}-{bc})/{sp})'
        ws[f'{C["Price Flag"]}{r}'] = (
            f'=IF({bc}="","No Data",IF({bc}>{lc},"Cost Up",IF({bc}<{lc},"Cost Down","Unchanged")))'
        )
        ws[f'{C["Margin Flag"]}{r}'] = (
            f'=IF({bc}="","No Data",IF(({sp}-{bc})/{sp}<0,"LOSS",'
            f'IF(({sp}-{bc})/{sp}<{lm},"Low Margin","Healthy")))'
        )
        # Selling price above the best supplier cost -> worth buying in (import).
        # Below it -> the local market pays less than it costs to source, so the
        # stock is better sold out of market (export). Equal -> break-even.
        ws[f'{C["Trade Direction"]}{r}'] = (
            f'=IF({bc}="","",IF({sp}>{bc},"Good for Import",'
            f'IF({sp}<{bc},"Good for Export","Break-even")))'
        )

    # Replace blank supplier cells with "not carried"
    for code in codes:
        c_idx = col[code]
        for r in range(2, nrow + 1):
            cell = ws.cell(row=r, column=c_idx)
            if cell.value is None:
                cell.value = NOT_CARRIED

    total_cols = ws.max_column
    supplier_idx = {col[c] for c in codes}
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        if c in supplier_idx:
            cell.fill = supp_fill
        elif c >= start or cell.value in ("Best Cost", "Best Supplier"):
            cell.fill = calc_fill
        else:
            cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32

    money_names = ["Last Cost", "Market/Selling Price", "Best Cost",
                   "Cost Diff (AED)", "Margin"] + codes
    pct_names = ["Cost Change %", "Margin %"]
    money_cols = {col[n] for n in money_names if n in col} | \
                 {column_index_from_string(C[n]) for n in money_names if n in C}
    pct_cols = {column_index_from_string(C[n]) for n in pct_names if n in C}

    grey = Font(name=FONT, size=10, color="9AA0AB", italic=True)
    for r in range(2, nrow + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c in supplier_idx and cell.value == NOT_CARRIED:
                cell.font = grey
                cell.alignment = Alignment(horizontal="center")
                continue
            cell.font = Font(name=FONT, size=10)
            if c in money_cols:
                cell.number_format = "#,##0.00"
            if c in pct_cols:
                cell.number_format = "0.0%"

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    widths = {"Barcode": 13, "SKU": 10, "Brand": 12, "Perfume Name": 20, "Description": 28,
              "Size (ml)": 8, "Gender": 8, "Category": 9, "Stock Qty": 8,
              "Country of Origin": 13, "Last Cost": 10, "Market/Selling Price": 13,
              "Best Cost": 11, "Best Supplier": 12, "Match Status": 20,
              "Cost Diff (AED)": 12, "Cost Change %": 11, "Margin": 10,
              "Margin %": 10, "Price Flag": 11, "Margin Flag": 12,
              "Trade Direction": 16}
    for i, cell in enumerate(ws[1]):
        ws.column_dimensions[get_column_letter(i + 1)].width = widths.get(cell.value, 12)

    # Flag colours
    red = PatternFill("solid", fgColor="F4CCCC")
    redf = Font(name=FONT, color="990000", size=10, bold=True)
    yel = PatternFill("solid", fgColor="FFF2CC")
    grn = PatternFill("solid", fgColor="D9EAD3")
    mf, pf = C["Margin Flag"], C["Price Flag"]
    ws.conditional_formatting.add(f"{mf}2:{mf}{nrow}",
        CellIsRule(operator="equal", formula=['"LOSS"'], fill=red, font=redf))
    ws.conditional_formatting.add(f"{mf}2:{mf}{nrow}",
        CellIsRule(operator="equal", formula=['"Low Margin"'], fill=yel))
    ws.conditional_formatting.add(f"{mf}2:{mf}{nrow}",
        CellIsRule(operator="equal", formula=['"Healthy"'], fill=grn))
    ws.conditional_formatting.add(f"{pf}2:{pf}{nrow}",
        CellIsRule(operator="equal", formula=['"Cost Up"'], fill=red))
    ws.conditional_formatting.add(f"{pf}2:{pf}{nrow}",
        CellIsRule(operator="equal", formula=['"Cost Down"'], fill=grn))

    # Trade Direction: import (buy in) green, export (sell out) amber
    td = C["Trade Direction"]
    blue = PatternFill("solid", fgColor="DEEAF6")
    amber_f = PatternFill("solid", fgColor="FCE4D6")
    ws.conditional_formatting.add(f"{td}2:{td}{nrow}",
        CellIsRule(operator="equal", formula=['"Good for Import"'], fill=grn))
    ws.conditional_formatting.add(f"{td}2:{td}{nrow}",
        CellIsRule(operator="equal", formula=['"Good for Export"'], fill=amber_f))
    ws.conditional_formatting.add(f"{td}2:{td}{nrow}",
        CellIsRule(operator="equal", formula=['"Break-even"'], fill=blue))

    # Highlight the winning supplier's code
    bs = get_column_letter(col["Best Supplier"])
    ws.conditional_formatting.add(f"{bs}2:{bs}{nrow}",
        CellIsRule(operator="notEqual", formula=['""'], fill=grn))

    _add_multi_summary(wb, nrow, codes, per_supplier, C, bs, mf)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _add_multi_summary(wb, nrow, codes, per_supplier, C, bs, mf):
    td = C["Trade Direction"]
    s = wb.create_sheet("Summary", 0)
    s.sheet_view.showGridLines = False
    t = s.cell(row=1, column=1, value="Multi-Supplier Cost & Margin Comparison")
    t.font = Font(name=FONT, bold=True, size=14, color="1F3864")
    s.cell(row=2, column=1,
           value="Each product priced across all suppliers. Best Cost = cheapest available.").font = \
        Font(name=FONT, italic=True, size=9, color="808080")

    r = 4
    s.cell(row=r, column=1, value="Sheets used").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for code, name, ref, _ in per_supplier:
        s.cell(row=r, column=1, value=f"{code} \u2014 {name}").font = Font(name=FONT, size=10)
        s.cell(row=r, column=2, value=ref).font = Font(name=FONT, size=10, color="6B7280")
        r += 1

    r += 1
    s.cell(row=r, column=1, value="Where the best price comes from").font = Font(name=FONT, bold=True, size=11)
    r += 1
    for code, name, ref, _ in per_supplier:
        s.cell(row=r, column=1, value=f"{code} cheapest on").font = Font(name=FONT, size=10, bold=True)
        s.cell(row=r, column=2,
               value=f'=COUNTIF(Comparison!{bs}2:{bs}{nrow},"{code}")').font = Font(name=FONT, size=10)
        r += 1

    r += 1
    rows = [
        ("Total products", f"=COUNTA(Comparison!A2:A{nrow})"),
        ("Available from a supplier", f'=COUNTIF(Comparison!{bs}2:{bs}{nrow},"?*")'),
        ("Items at a LOSS", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"LOSS")'),
        ("Items at Low Margin", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"Low Margin")'),
        ("Items Healthy margin", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"Healthy")'),
        ("Good for Import", f'=COUNTIF(Comparison!{td}2:{td}{nrow},"Good for Import")'),
        ("Good for Export", f'=COUNTIF(Comparison!{td}2:{td}{nrow},"Good for Export")'),
        ("Avg margin % (best cost)", f'=AVERAGE(Comparison!{C["Margin %"]}2:{C["Margin %"]}{nrow})'),
    ]
    for label, formula in rows:
        lab = s.cell(row=r, column=1, value=label)
        lab.font = Font(name=FONT, size=10, bold=True)
        val = s.cell(row=r, column=2, value=formula)
        val.font = Font(name=FONT, size=10)
        if "%" in label:
            val.number_format = "0.0%"
        if "LOSS" in label:
            lab.fill = PatternFill("solid", fgColor="F4CCCC")
            val.fill = PatternFill("solid", fgColor="F4CCCC")
        elif "Low Margin" in label:
            lab.fill = PatternFill("solid", fgColor="FFF2CC")
            val.fill = PatternFill("solid", fgColor="FFF2CC")
        r += 1

    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 20
