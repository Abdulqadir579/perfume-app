"""
Core comparison engine.
Matches a shop master sheet against a supplier sheet by Barcode and produces
a formatted comparison workbook with cost-difference, margin, and flag columns.

This is the SAME logic validated in the Excel demo. No simulation here —
it reads the supplier's real Current Cost.
"""
from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

FONT = "Arial"
LOW_MARGIN_THRESHOLD = 0.15  # below this = "Low Margin"
DELTA = "\u0394"  # Greek delta, kept out of f-string expressions for py3.9 compatibility

# Columns we require to exist (case-insensitive, trimmed)
MASTER_REQUIRED = ["Barcode", "Last Cost", "Market/Selling Price"]
SUPPLIER_REQUIRED = ["Barcode", "Current Cost"]

OUTPUT_COLS = [
    "Barcode", "SKU", "Brand", "Perfume Name", "Description", "Size (ml)",
    "Gender", "Category", "Stock Qty", "Country of Origin",
    "Last Cost", "Supplier Current Cost", "Market/Selling Price", "Match Status",
]


class CompareError(Exception):
    """Raised for user-fixable problems (missing columns, unreadable file)."""


def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _check_columns(df: pd.DataFrame, required, label):
    have = {c.lower(): c for c in df.columns}
    missing = [c for c in required if c.lower() not in have]
    if missing:
        raise CompareError(
            f"The {label} file is missing required column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(df.columns)}"
        )
    # return actual-cased names
    return {c: have[c.lower()] for c in required}


def build_comparison(master_bytes: bytes, supplier_bytes: bytes) -> BytesIO:
    """Take two uploaded xlsx files (as bytes), return a formatted xlsx in a BytesIO."""
    try:
        master = _norm_cols(pd.read_excel(BytesIO(master_bytes)))
    except Exception as e:
        raise CompareError(f"Could not read the shop master file as Excel. ({e})")
    try:
        supplier = _norm_cols(pd.read_excel(BytesIO(supplier_bytes)))
    except Exception as e:
        raise CompareError(f"Could not read the supplier file as Excel. ({e})")

    m_cols = _check_columns(master, MASTER_REQUIRED, "shop master")
    s_cols = _check_columns(supplier, SUPPLIER_REQUIRED, "supplier")

    # Pull only the unique key + live price from supplier
    supp_slim = supplier[[s_cols["Barcode"], s_cols["Current Cost"]]].rename(
        columns={s_cols["Barcode"]: "Barcode", s_cols["Current Cost"]: "Supplier Current Cost"}
    )
    # De-dupe supplier on barcode (keep last occurrence) to avoid row explosion
    supp_slim = supp_slim.drop_duplicates(subset="Barcode", keep="last")

    shop = master.rename(columns={m_cols["Barcode"]: "Barcode"})
    if "Current Cost" in shop.columns:
        shop = shop.drop(columns=["Current Cost"])  # drop stale shop copy; supplier provides live

    out = shop.merge(supp_slim, on="Barcode", how="left")
    out["Match Status"] = np.where(
        out["Supplier Current Cost"].notna(), "Matched (Barcode)", "NOT FOUND in supplier sheet"
    )

    # Keep only known output columns that exist (be tolerant of missing optional ones)
    present = [c for c in OUTPUT_COLS if c in out.columns]
    out = out[present]

    return _style_workbook(out)


def _style_workbook(df: pd.DataFrame) -> BytesIO:
    buf0 = BytesIO()
    df.to_excel(buf0, index=False)
    buf0.seek(0)
    wb = load_workbook(buf0)
    ws = wb.active
    ws.title = "Comparison"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    calc_hdr_fill = PatternFill("solid", fgColor="C55A11")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    nrow = ws.max_row
    ncol = ws.max_column
    col = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    last_cost_c = get_column_letter(col["Last Cost"])
    supp_cost_c = get_column_letter(col["Supplier Current Cost"])
    sell_c = get_column_letter(col["Market/Selling Price"])

    new_headers = ["Cost Diff (AED)", "Cost Change %", "Old Margin", "New Margin",
                   "Old Margin %", "New Margin %", "Margin " + DELTA + " (pts)",
                   "Price Flag", "Margin Flag"]
    start = ncol + 1
    for i, h in enumerate(new_headers):
        ws.cell(row=1, column=start + i, value=h)
    C = {h: get_column_letter(start + i) for i, h in enumerate(new_headers)}

    lm = LOW_MARGIN_THRESHOLD
    _md = "Margin " + DELTA + " (pts)"
    for r in range(2, nrow + 1):
        lc, sc, sp = f"{last_cost_c}{r}", f"{supp_cost_c}{r}", f"{sell_c}{r}"
        ws[f'{C["Cost Diff (AED)"]}{r}'] = f'=IF({sc}="","",{sc}-{lc})'
        ws[f'{C["Cost Change %"]}{r}'] = f'=IF(OR({sc}="",{lc}=0),"",({sc}-{lc})/{lc})'
        ws[f'{C["Old Margin"]}{r}'] = f'={sp}-{lc}'
        ws[f'{C["New Margin"]}{r}'] = f'=IF({sc}="","",{sp}-{sc})'
        ws[f'{C["Old Margin %"]}{r}'] = f'=IF({sp}=0,"",({sp}-{lc})/{sp})'
        ws[f'{C["New Margin %"]}{r}'] = f'=IF(OR({sc}="",{sp}=0),"",({sp}-{sc})/{sp})'
        ws[C[_md] + str(r)] = f'=IF(OR({sc}="",{sp}=0),"",(({sp}-{sc})/{sp})-(({sp}-{lc})/{sp}))'
        ws[f'{C["Price Flag"]}{r}'] = f'=IF({sc}="","No Data",IF({sc}>{lc},"Cost Up",IF({sc}<{lc},"Cost Down","Unchanged")))'
        ws[f'{C["Margin Flag"]}{r}'] = f'=IF({sc}="","No Data",IF(({sp}-{sc})/{sp}<0,"LOSS",IF(({sp}-{sc})/{sp}<{lm},"Low Margin","Healthy")))'

    total_cols = ws.max_column
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = calc_hdr_fill if c >= start else hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    from openpyxl.utils import column_index_from_string
    money_names = ["Last Cost", "Supplier Current Cost", "Market/Selling Price",
                   "Cost Diff (AED)", "Old Margin", "New Margin"]
    pct_names = ["Cost Change %", "Old Margin %", "New Margin %", "Margin " + DELTA + " (pts)"]
    money_cols = {col[n] for n in money_names if n in col} | \
                 {column_index_from_string(C[n]) for n in money_names if n in C}
    pct_cols = {column_index_from_string(C[n]) for n in pct_names if n in C}
    for r in range(2, nrow + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = border
            if c in money_cols:
                cell.number_format = "#,##0.00"
            if c in pct_cols:
                cell.number_format = "0.0%"

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    widths = {"Barcode": 13, "SKU": 10, "Brand": 12, "Perfume Name": 20, "Description": 30,
              "Size (ml)": 8, "Gender": 8, "Category": 9, "Stock Qty": 8, "Country of Origin": 13,
              "Last Cost": 10, "Supplier Current Cost": 14, "Market/Selling Price": 13,
              "Match Status": 18, "Cost Diff (AED)": 12, "Cost Change %": 11,
              "Old Margin": 11, "New Margin": 11, "Old Margin %": 11,
              "New Margin %": 11, "Margin \u0394 (pts)": 12, "Price Flag": 11, "Margin Flag": 12}
    for i, cell in enumerate(ws[1]):
        ws.column_dimensions[get_column_letter(i + 1)].width = widths.get(cell.value, 12)

    red = PatternFill("solid", fgColor="F4CCCC")
    redf = Font(name=FONT, color="990000", size=10, bold=True)
    yel = PatternFill("solid", fgColor="FFF2CC")
    grn = PatternFill("solid", fgColor="D9EAD3")
    mf, pf = C["Margin Flag"], C["Price Flag"]
    rng_mf = f"{mf}2:{mf}{nrow}"
    ws.conditional_formatting.add(rng_mf, CellIsRule(operator="equal", formula=['"LOSS"'], fill=red, font=redf))
    ws.conditional_formatting.add(rng_mf, CellIsRule(operator="equal", formula=['"Low Margin"'], fill=yel))
    ws.conditional_formatting.add(rng_mf, CellIsRule(operator="equal", formula=['"Healthy"'], fill=grn))
    rng_pf = f"{pf}2:{pf}{nrow}"
    ws.conditional_formatting.add(rng_pf, CellIsRule(operator="equal", formula=['"Cost Up"'], fill=red))
    ws.conditional_formatting.add(rng_pf, CellIsRule(operator="equal", formula=['"Cost Down"'], fill=grn))
    ms = get_column_letter(col["Match Status"])
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(total_cols)}{nrow}",
        FormulaRule(formula=[f'NOT(ISNUMBER(SEARCH("Matched",${ms}2)))'],
                    fill=PatternFill("solid", fgColor="FCE4D6")),
    )

    _add_summary(wb, ws, nrow, ms, pf, mf, C)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _add_summary(wb, ws, nrow, ms, pf, mf, C):
    s = wb.create_sheet("Summary", 0)
    s.sheet_view.showGridLines = False
    t = s.cell(row=1, column=1, value="Perfume Cost & Margin Comparison \u2014 Summary")
    t.font = Font(name=FONT, bold=True, size=14, color="1F3864")
    s.cell(row=2, column=1, value="Matched by Barcode between shop master and supplier sheet.").font = \
        Font(name=FONT, italic=True, size=9, color="808080")

    rows = [
        ("Total products", f"=COUNTA(Comparison!A2:A{nrow})"),
        ("Matched with supplier", f'=COUNTIF(Comparison!{ms}2:{ms}{nrow},"Matched*")'),
        ("Not found in supplier", f'=COUNTIF(Comparison!{ms}2:{ms}{nrow},"NOT*")'),
        ("Cost increased", f'=COUNTIF(Comparison!{pf}2:{pf}{nrow},"Cost Up")'),
        ("Cost decreased", f'=COUNTIF(Comparison!{pf}2:{pf}{nrow},"Cost Down")'),
        ("Items now at a LOSS", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"LOSS")'),
        ("Items at Low Margin", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"Low Margin")'),
        ("Items Healthy margin", f'=COUNTIF(Comparison!{mf}2:{mf}{nrow},"Healthy")'),
        ("Avg new margin %", f'=AVERAGE(Comparison!{C["New Margin %"]}2:{C["New Margin %"]}{nrow})'),
        ("Avg cost change %", f'=AVERAGE(Comparison!{C["Cost Change %"]}2:{C["Cost Change %"]}{nrow})'),
    ]
    r0 = 4
    for i, (label, formula) in enumerate(rows):
        rr = r0 + i
        lc = s.cell(row=rr, column=1, value=label)
        lc.font = Font(name=FONT, size=10, bold=True)
        vc = s.cell(row=rr, column=2, value=formula)
        vc.font = Font(name=FONT, size=10)
        if "%" in label:
            vc.number_format = "0.0%"
        fill = "F4CCCC" if "LOSS" in label else ("FFF2CC" if "Low Margin" in label else None)
        if fill:
            lc.fill = PatternFill("solid", fgColor=fill)
            vc.fill = PatternFill("solid", fgColor=fill)
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 16
